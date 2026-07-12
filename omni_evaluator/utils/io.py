# OmniEvaluator
# Copyright (c) 2026-present NAVER Cloud Corp.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import csv
import gzip
import io
import json
import logging
import math
import numpy as np
from omegaconf import ListConfig, DictConfig
import openpyxl
import os
from pathlib import Path, PosixPath
import pickle
import PIL
from PIL import Image
import soundfile
import tempfile
from typing import Any, Dict, List, Union, Tuple, Optional, Callable, Iterable
from urllib.parse import urlparse, urlunparse
import yaml

logger = logging.getLogger(__name__)

# Keys copied from legacy run_outputs entries into per-run records
_LEGACY_RUN_OUTPUT_KEYS = [
    "prediction", "prediction_postprocessed", "reasoning_content",
    "tool_calls", "perplexities", "latency", "metrics",
]


def resolve_gguf_path(model_name_or_path: str, filename: str) -> Optional[str]:
    """Resolve a local ``.gguf`` file path: a direct file, or the first match of the
    ``filename`` glob inside a directory. Returns None when ``model_name_or_path`` is a
    HF hub repo id (no local file), so the caller can fall back to a hub download."""
    from glob import glob
    if os.path.isfile(model_name_or_path):
        return model_name_or_path
    if os.path.isdir(model_name_or_path):
        matches = sorted(glob(os.path.join(model_name_or_path, filename)))
        if not matches:
            raise FileNotFoundError(
                f"no GGUF matching {filename!r} under {model_name_or_path!r}")
        return matches[0]
    return None


def ensure_per_run_format(inference_data: list) -> list:
    """Convert legacy flat-records-with-run_outputs to List[List[Dict]].

    Old format: List[Dict] where each dict has a ``run_outputs`` list.
    New format: List[List[Dict]] — one list of records per run.

    If the data already looks like the new format (first element is a list)
    it is returned unchanged.
    """
    if not inference_data or isinstance(inference_data[0], list):
        return inference_data

    num_runs = len(inference_data[0].get("run_outputs", [])) if inference_data else 1
    if num_runs < 1:
        num_runs = 1

    per_run: list = [[] for _ in range(num_runs)]
    for record in inference_data:
        run_outputs = record.get("run_outputs", [])
        base = {k: v for k, v in record.items() if k not in ("run_outputs", "num_runs")}
        for ri in range(num_runs):
            run_record = dict(base)
            if ri < len(run_outputs):
                for key in _LEGACY_RUN_OUTPUT_KEYS:
                    if key in run_outputs[ri]:
                        run_record[key] = run_outputs[ri][key]
            per_run[ri].append(run_record)
    return per_run


class _NonSerializableEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle non-serializable objects like callables."""
    def default(self, obj):
        if callable(obj):
            return f"<callable: {getattr(obj, '__name__', repr(obj))}>"
        if isinstance(obj, (np.integer, np.floating)):
            return float(obj) if isinstance(obj, np.floating) else int(obj)
        try:
            return super().default(obj)
        except TypeError:
            return str(obj)


def _json_safe(obj: Any) -> Any:
    """Recursively replace non-finite floats (NaN / ±Infinity) with ``None``.

    Python's ``json.dump`` writes ``NaN``/``Infinity`` as literal tokens
    (``allow_nan=True`` default), which are NOT valid JSON — strict parsers
    (ijson, JS ``JSON.parse``, ``jq``, Go/Rust) reject them, and ``NaN`` also
    conflates "not computed" with a real value. We emit ``null`` instead so the
    output round-trips everywhere. ``json.JSONEncoder.default`` cannot intercept
    native floats (they are "serializable"), so we sanitize before dumping.

    Container nodes are rebuilt but leaf strings/ints are returned by reference,
    so this is cheap even for large generation outputs (no string copies).
    """
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else None
    if isinstance(obj, np.floating):
        _f = float(obj)
        return _f if math.isfinite(_f) else None
    if isinstance(obj, dict):
        return {_k: _json_safe(_v) for _k, _v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(_v) for _v in obj]
    return obj


def read_file(
    filepath: Union[str, PosixPath],
    encoding: str = "utf-8",
) -> Any:
    """Read and deserialize a file based on its extension (json, jsonl, yaml, csv, etc.)."""
    if isinstance(filepath, PosixPath):
        filepath = Path.as_posix(filepath)
    if not (isinstance(filepath, str) and os.path.exists(filepath)):
        raise ValueError(f'Invalid filepath or not exist: {filepath}')

    data = None
    if filepath.endswith(".json"):
        with open(filepath, "r", encoding="utf-8") as fp:
            data = json.load(fp)
    elif filepath.endswith(".jsonl"):
        data = list()
        with open(filepath, "r", encoding="utf-8") as fp:
            for row in fp:
                data.append(json.loads(row.strip()))
    elif filepath.endswith(".pickle"):
        logger.warning(
            f"Loading pickle file: {filepath}. "
            "Pickle deserialization can execute arbitrary code. "
            "Only load pickle files from trusted sources."
        )
        with open(filepath, "rb") as fp:
            data = pickle.load(fp)
    elif filepath.endswith(".gzip"):
        data = list()
        with gzip.open(filepath, "rt", encoding=encoding) as fp:
            for row in fp:
                data.append(json.loads(row))
    elif (
        filepath.endswith(".csv")
        or filepath.endswith(".tsv")
    ):
        delimiter = ","
        if filepath.endswith(".tsv"):
            delimiter = "\t"
        data = list()
        with open(filepath, newline="", encoding=encoding) as fp:
            reader = csv.reader(fp, delimiter=delimiter)
            data = list(reader)
    elif filepath.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(filename=filepath, data_only=True)
        data = list()
        for row in workbook.active.iter_rows(values_only=True): # iter first sheet\
            data.append(list(row))
    elif filepath.endswith(".yaml"):
        with open(filepath, "rb") as fp:
            data = yaml.safe_load(fp)
    elif filepath.endswith(".wav"):
        data, _sampling_rate = soundfile.read(filepath)
    else:
        with open(filepath, "rb") as fp:
            data = fp.read()
        # raise ValueError(f'# [error] unsupported file type: {Path(filepath).suffix}')
    return data


def iter_file(
    filepath: Union[str, PosixPath],
    encoding: str = "utf-8",
) -> Iterable[Any]:
    """Lazily iterate row-oriented files (jsonl, gzip-jsonl, csv, tsv).

    Yields one parsed row at a time instead of materializing the whole file in
    memory as ``read_file`` does. Use this for large dataset files where the
    eager list would block the first inference for tens of seconds. Caller is
    responsible for separately obtaining the size (e.g. via ``count_lines``).
    """
    if isinstance(filepath, PosixPath):
        filepath = Path.as_posix(filepath)
    if not (isinstance(filepath, str) and os.path.exists(filepath)):
        raise ValueError(f'Invalid filepath or not exist: {filepath}')

    if filepath.endswith(".jsonl"):
        with open(filepath, "r", encoding=encoding) as fp:
            for row in fp:
                yield json.loads(row.strip())
    elif filepath.endswith(".gzip"):
        with gzip.open(filepath, "rt", encoding=encoding) as fp:
            for row in fp:
                yield json.loads(row)
    elif filepath.endswith(".csv") or filepath.endswith(".tsv"):
        delimiter = "\t" if filepath.endswith(".tsv") else ","
        with open(filepath, newline="", encoding=encoding) as fp:
            yield from csv.reader(fp, delimiter=delimiter)
    else:
        raise ValueError(
            f'iter_file: unsupported extension for lazy read: {filepath} '
            f'(supported: .jsonl, .gzip, .csv, .tsv). Use read_file for other types.'
        )


def count_lines(
    filepath: Union[str, PosixPath],
    encoding: str = "utf-8",
) -> int:
    """Return the number of lines in a row-oriented file without parsing rows.

    Used to pre-compute ``dataset_size`` for the lazy ``iter_file`` path; runs
    10-100x faster than ``read_file`` since no JSON/CSV parsing happens.
    """
    if isinstance(filepath, PosixPath):
        filepath = Path.as_posix(filepath)
    if not (isinstance(filepath, str) and os.path.exists(filepath)):
        raise ValueError(f'Invalid filepath or not exist: {filepath}')

    if filepath.endswith(".gzip"):
        with gzip.open(filepath, "rt", encoding=encoding) as fp:
            return sum(1 for _ in fp)
    with open(filepath, "r", encoding=encoding) as fp:
        return sum(1 for _ in fp)


def write_file(
    filepath: Union[str, PosixPath],
    obj: Any,
    encoding: str = "utf-8",
) -> str:
    """Serialize and write *obj* to *filepath* based on its extension. Returns the filepath."""
    if isinstance(filepath, PosixPath):
        filepath = Path.as_posix(filepath)
    if not Path(filepath).parent.exists():
        Path(filepath).parent.mkdir(exist_ok=True, parents=True)

    if filepath.endswith(".json"):
        with open(filepath, "w", encoding=encoding) as fp:
            json.dump(_json_safe(obj), fp, ensure_ascii=False, cls=_NonSerializableEncoder)
    elif filepath.endswith(".jsonl"):
        if not isinstance(obj, (list, ListConfig)):
            raise TypeError(f'Only list type can be written in .jsonl format')
        with open(filepath, "w", encoding=encoding) as fp:
            for row in obj:
                fp.write(json.dumps(_json_safe(row), ensure_ascii=False, cls=_NonSerializableEncoder)+"\n")
    elif filepath.endswith(".pickle"):
        with open(filepath, "wb") as fp:
            pickle.dump(obj, fp)
    elif (
        filepath.endswith(".csv")
        or filepath.endswith(".tsv")
    ):
        if not isinstance(obj, (list, ListConfig)):
            raise TypeError(f'Only list type can be written in .csv or .tsv format')
        delimiter = ","
        if filepath.endswith(".tsv"):
            delimiter = "\t"
        data = list()
        with open(filepath, mode="w", newline="", encoding=encoding) as fp:
            writer = csv.writer(fp, delimiter=delimiter)
            writer.writerows(obj)
    elif filepath.endswith(".xlsx"):
        if not isinstance(obj, (list, ListConfig)):
            raise TypeError(f'Only list type can be written in .xlsx format')
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for row in obj:
            worksheet.append(row)
        workbook.save(filepath)
    elif filepath.endswith(".wav"):
        _data, _sampling_rate = soundfile.read(io.BytesIO(obj))
        soundfile.write(filepath, _data, _sampling_rate)
    elif isinstance(obj, bytes):
        with open(filepath, "wb") as fp:
            fp.write(obj)
    else:
        raise ValueError(f'unsupported file type: {Path(filepath).suffix}')
    return filepath


def get_output_dirpath(
    output_dirpath: str,
    evaluation_engine: str,
    exp_name: str,
    version_name: Optional[str] = None,
) -> str:
    """Build the output directory path from experiment and engine names."""
    if (
        not isinstance(version_name, str)
        or len(version_name) < 1
    ):
        version_name = "checkpoint-none"
        
    output_dirpath = os.path.join(
        output_dirpath, 
        exp_name, 
        version_name,
    ) # e.g. ~/evaluator/v1.0/test_model/checkpoint-none
    output_dirpath = os.path.join(
        output_dirpath, 
        evaluation_engine,
    ) # e.g. ~/evaluator/v1.0/test_model/checkpoint-none/lm_eval_harness
    return output_dirpath


def get_output_filename(
    benchmark: str,
    evaluation_method: str,
) -> str:
    """Return the output filename in the form ``{benchmark}__{method}.json``."""
    output_filename = f"{benchmark}__{evaluation_method}"
    return f"{output_filename}.json"


def get_temp_filepath(
    prefix: Optional[str] = None,
    suffix: Optional[str] = None,
    dirpath: Optional[str] = None,
) -> str:
    """Create a temporary file and return its path."""
    if isinstance(dirpath, str) and not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)
    fd, filename = tempfile.mkstemp(
        suffix=suffix,
        prefix=prefix,
        dir=dirpath,
    )
    os.close(fd)
    return filename


def is_sub_path(
    parent_path: str,
    child_path: str, 
) -> bool:
    child_path = Path(child_path).resolve()
    parent_path = Path(parent_path).resolve()
    try:
        child_path.relative_to(parent_path)
        return True
    except ValueError as ex:
        return False